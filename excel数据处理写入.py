# -*- coding=utf-8 -*-

import xlrd
from xlrd import xldate_as_tuple
import datetime
import re
from openpyxl import load_workbook

'''
xlrd中单元格的数据类型
数字一律按浮点型输出，日期输出成一串小数，布尔型输出0或1，所以我们必须在程序中做判断处理转换
成我们想要的数据类型
0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
'''


class workExcel:

    def __init__(self, excel_path, sheet_name):
        self.excel_book = xlrd.open_workbook(r'%s' % excel_path)  # 读取excel，路径使用原生字符串
        self.excel_table = self.excel_book.sheet_by_name(sheet_name)  # 根据sheet名称读取
        print('Sheet读取检查：', self.excel_book.sheet_loaded(sheet_name))
        self.keys = self.excel_table.row_values(0)  # 读取行标题
        print('标题栏：', self.keys)
        self.nrows = self.excel_table.nrows  # 读取行
        self.ncols = self.excel_table.ncols  # 读取列
        print('行数：', self.nrows)
        print('列数：', self.ncols)

    def readData(self):     # 将表格内的数据读取为列表
        # 定义一个空列表
        table_data = []
        for i in range(1, self.nrows):
            # 定义一个空字典
            sheet_data = {}
            for j in range(self.ncols):
                # 获取单元格数据类型
                f_type = self.excel_table.cell(i, j).ctype
                # 获取单元格数据
                f_cell = self.excel_table.cell_value(i, j)
                if f_type == 2 and f_cell % 1 == 0:  # 如果是整形
                    f_cell = int(f_cell)
                elif f_type == 3:
                    # 转成datetime对象
                    date = datetime.datetime(*xldate_as_tuple(f_cell, 0))
                    f_cell = date.strftime('%Y-%m-%d')  # /%m %H:%M:%S
                elif f_type == 4:
                    f_cell = True if f_cell == 1 else False
                sheet_data[self.keys[j]] = f_cell
                # 循环每一个有效的单元格，将字段与值对应存储到字典中
                # 字典的key就是excel表中每列第一行的字段
                # sheet_data[self.keys[j]] = self.table.row_values(i)[j]
            # 再将字典追加到列表中
            table_data.append(sheet_data)
        # 返回从excel中获取到的数据：以列表存字典的形式返回
        return table_data

    def fxData_1(self):     # 以下五个函数对应五个公式
        unpd_data = self.readData()
        fx_1 = []
        for i in range(self.nrows - 1):
            # print(unpd_data[i].get('date'))
            open_data = unpd_data[i].get('OPEN')
            date = unpd_data[i].get('date')
            month = re.findall(r'-(.*)-', date)
            # print('断点打印数据：', month)
            if month == ['02']:
                fx_1.append(open_data)
            else:
                data_minus = open_data - unpd_data[i - 1].get('OPEN')
                fx_1.append(data_minus)
        return fx_1

    def fxData_2(self):
        unpd_data = self.readData()
        fx_2 = []
        for i in range(self.nrows - 1):
            sale_data = unpd_data[i].get('SALE')
            date = unpd_data[i].get('date')
            month = re.findall(r'-(.*)-', date)
            # print('断点打印数据：', month)
            if month == ['02']:
                fx_2.append(sale_data)
            else:
                data_minus = sale_data - unpd_data[i - 1].get('SALE')
                fx_2.append(data_minus)
        return fx_2

    def fxData_3(self):
        unpd_data = self.fxData_1()
        fx_3 = []
        for i in range(11, len(self.readData())):
            data_fx3 = unpd_data[i] / unpd_data[i - 11] - 1
            fx_3.append(data_fx3)
        return fx_3

    def fxData_4(self):
        unpd_fx1 = self.fxData_1()
        unpd_fx2 = self.fxData_2()
        fx_4 = []
        data_fx4 = unpd_fx1[0] - unpd_fx2[0]
        fx_4.append(data_fx4)
        for i in range(1, self.nrows - 1):
            data_fx4 = data_fx4 + unpd_fx1[i] - unpd_fx2[i]
            fx_4.append(data_fx4)
        return fx_4

    def fxData_5(self):
        unpd_fx4 = self.fxData_4()
        unpd_fx2 = self.fxData_2()
        fx_5 = []
        for i in range(self.nrows - 1):
            data_fx5 = unpd_fx4[i] / unpd_fx2[i]
            fx_5.append(data_fx5)
        return fx_5

    def writeData(self):    # 写入到文件
        wb = load_workbook("% s" % excel_path)  # 生成一个已存在的workbook对象
        wb_active = wb.active  # 激活sheet
        wb_active.cell(1, 4, '数据1')
        wb_active.cell(1, 5, '数据2')
        wb_active.cell(1, 6, '数据3')
        wb_active.cell(1, 7, '数据4')
        wb_active.cell(1, 8, '数据5')
        fx_1 = self.fxData_1()
        fx_2 = self.fxData_2()
        fx_3 = self.fxData_3()
        fx_4 = self.fxData_4()
        fx_5 = self.fxData_5()
        for i in range(self.nrows - 1):
            wb_active.cell(i+2, 4, fx_1[i])     # 往sheet中的第i+2行第4列写入列表中的数据
            wb_active.cell(i+2, 5, fx_2[i])
            wb_active.cell(i+2, 7, fx_4[i])
            wb_active.cell(i+2, 8, fx_5[i])

        for j in range(len(self.fxData_3())):
            wb_active.cell(j+13, 6, fx_3[j])

        wb.save("% s" % excel_path)  # 保存
        print('写入excel操作完成！')


if __name__ == '__main__':
    excel_path = r"./DATA.xlsx"     # 设置excel文件路径
    sheet_name = "Sheet1"   # 输入sheet名称
    get_data = workExcel(excel_path, sheet_name)
    datas = get_data.readData()
    print('EXCEL读取列表：', datas)
    data1 = get_data.fxData_1()
    data2 = get_data.fxData_2()
    data3 = get_data.fxData_3()
    data4 = get_data.fxData_4()
    data5 = get_data.fxData_5()
    print('数据1：', data1)
    print('数据2：', data2)
    print('数据3：', data3)
    print('数据4：', data4)
    print('数据5：', data5)
    # get_data.writeData()    # 写入文件
